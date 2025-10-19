import os
import io
import json
import zipfile
from datetime import datetime
from flask import Flask, request, jsonify, send_file, render_template, redirect, url_for
import pandas as pd
import qrcode
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx'}
STATE_FILE = os.path.join(UPLOAD_FOLDER, 'state.json')

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Utility functions

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def save_state(state):
    with open(STATE_FILE, 'w', encoding='utf-8') as f:
        json.dump(state, f)


def load_state():
    if not os.path.exists(STATE_FILE):
        return {}
    with open(STATE_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload():
    # Upload an XLSX file containing the student list
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(path)

        # Read student data with pandas to infer columns
        df = pd.read_excel(path)
        # Expect at least two columns: Student ID and Name. If names differ, we keep first two columns.
        # Normalize column names for internal use
        cols = df.columns.tolist()
        if len(cols) < 2:
            return jsonify({'error': 'Spreadsheet must contain at least two columns (ID and Name)'}), 400

        # Save state
        state = {
            'workbook_path': path,
            'id_column': cols[0],
            'name_column': cols[1],
            'lectures': [],  # list of {label, code, col_letter}
            'current_lecture': None
        }
        save_state(state)
        return jsonify({'message': 'Uploaded', 'id_column': cols[0], 'name_column': cols[1]})
    return jsonify({'error': 'Invalid file'}), 400


@app.route('/create_lecture', methods=['POST'])
def create_lecture():
    # Body: {"label": "Lecture 1", "code": "ABC123"}
    payload = request.get_json(force=True)
    label = payload.get('label') or datetime.utcnow().isoformat()
    code = str(payload.get('code') or '')
    if code == '':
        return jsonify({'error': 'Lecture code required'}), 400

    state = load_state()
    wb_path = state.get('workbook_path')
    if not wb_path or not os.path.exists(wb_path):
        return jsonify({'error': 'No workbook uploaded'}), 400

    # Add new column to workbook for this lecture
    wb = load_workbook(wb_path)
    ws = wb.active

    # Append header in the next free column
    max_col = ws.max_column
    new_col_idx = max_col + 1
    header_cell = ws.cell(row=1, column=new_col_idx)
    header_cell.value = label
    wb.save(wb_path)

    lecture = {
        'label': label,
        'code': code,
        'col_idx': new_col_idx
    }
    state['lectures'].append(lecture)
    state['current_lecture'] = lecture
    save_state(state)

    return jsonify({'message': 'Lecture created', 'lecture': lecture})


@app.route('/set_current_lecture', methods=['POST'])
def set_current_lecture():
    payload = request.get_json(force=True)
    idx = payload.get('index')
    state = load_state()
    lectures = state.get('lectures', [])
    if idx is None or idx < 0 or idx >= len(lectures):
        return jsonify({'error': 'Invalid lecture index'}), 400
    state['current_lecture'] = lectures[idx]
    save_state(state)
    return jsonify({'message': 'Current lecture set', 'current': state['current_lecture']})


@app.route('/generate_qr', methods=['GET'])
def generate_qr():
    # Generate a ZIP with one QR per student for the current lecture
    state = load_state()
    wb_path = state.get('workbook_path')
    lecture = state.get('current_lecture')
    if not lecture:
        return jsonify({'error': 'No current lecture set'}), 400
    if not wb_path or not os.path.exists(wb_path):
        return jsonify({'error': 'No workbook uploaded'}), 400

    df = pd.read_excel(wb_path)
    id_col = state['id_column']
    name_col = state['name_column']

    mem_zip = io.BytesIO()
    with zipfile.ZipFile(mem_zip, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        for _, row in df.iterrows():
            sid = str(row[id_col])
            name = str(row[name_col])
            payload = json.dumps({'id': sid, 'name': name, 'lecture_code': lecture['code']})
            img = qrcode.make(payload)
            img_byte_arr = io.BytesIO()
            img.save(img_byte_arr, format='PNG')
            img_byte_arr.seek(0)
            fname = f"{sid}_{secure_filename(name)}.png"
            zf.writestr(fname, img_byte_arr.read())

    mem_zip.seek(0)
    return send_file(mem_zip, download_name='qrcodes.zip', as_attachment=True)


@app.route('/scan', methods=['POST'])
def scan():
    # Called by frontend after scanning — body: {"raw": "..."}
    data = request.get_json(force=True)
    raw = data.get('raw')
    try:
        payload = json.loads(raw)
    except Exception:
        return jsonify({'error': 'Invalid QR payload'}), 400

    sid = str(payload.get('id'))
    embedded_code = str(payload.get('lecture_code'))

    state = load_state()
    lecture = state.get('current_lecture')
    if not lecture:
        return jsonify({'error': 'No active lecture'}), 400

    if embedded_code != str(lecture['code']):
        return jsonify({'error': 'Lecture code mismatch'}), 403

    # Mark attendance in workbook by placing a check mark in the lecture's column
    wb_path = state.get('workbook_path')
    wb = load_workbook(wb_path)
    ws = wb.active

    id_col_name = state['id_column']
    # Find header indices to map DataFrame columns to Excel columns
    # We'll iterate rows to find matching student id in first column that corresponds to id_col index
    # Simpler assumption: student ID is in column 1 (A). If not, try to find the column by reading header row.

    # Find id column index by searching header row for id_col_name
    id_col_idx = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == id_col_name:
            id_col_idx = c
            break
    if id_col_idx is None:
        id_col_idx = 1

    target_col = lecture['col_idx']

    # Find the row for this student
    target_row = None
    for r in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=id_col_idx).value
        if cell_val is None:
            continue
        if str(cell_val) == sid:
            target_row = r
            break

    if target_row is None:
        return jsonify({'error': 'Student ID not found'}), 404

    ws.cell(row=target_row, column=target_col).value = '✓'
    wb.save(wb_path)

    return jsonify({'message': 'Attendance recorded', 'student': sid})


@app.route('/export', methods=['GET'])
def export():
    state = load_state()
    wb_path = state.get('workbook_path')
    if not wb_path or not os.path.exists(wb_path):
        return jsonify({'error': 'No workbook uploaded'}), 400
    return send_file(wb_path, as_attachment=True)


if __name__ == '__main__':
    app.run(debug=True)
